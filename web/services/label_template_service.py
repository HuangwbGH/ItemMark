import os
import re
from typing import Any, Dict, List, Tuple

import openpyxl

from services.module_config_service import ModuleConfig, resolve_module_path


DEFAULTS: Dict[str, Any] = {
    "label_width_mm": 100,
    "label_min_height_mm": 58,
    "label_padding_mm": 4,
    "name_font_family": "Microsoft YaHei, PingFang SC, Arial, sans-serif",
    "name_font_size_pt": 13,
    "name_max_lines": 2,
    "field_font_family": "Microsoft YaHei, PingFang SC, Arial, sans-serif",
    "field_font_size_pt": 8.5,
    "remark_max_lines": 4,
    "field_label_width_px": 52,
    "qr_show": True,
    "qr_size_mm": 22,
    "labels_per_row": 2,
    "label_gap_mm": 6,
    "labels_per_page": 4,
    "footer_show": True,
    "footer_font_family": "Microsoft YaHei, PingFang SC, Arial, sans-serif",
    "footer_font_size_pt": 6,
    "footer_left_text": "{material_code}",
    "footer_right_text": "{print_time}",
}

DEFAULT_FIELDS: List[Dict[str, Any]] = [
    {"key": "material_code", "label": "物料编号", "show": True, "order": 1},
    {"key": "cloth_card_no", "label": "布卡号", "show": True, "order": 2},
    {"key": "specification", "label": "规格型号", "show": True, "order": 3},
    {"key": "color_desc", "label": "颜色", "show": True, "order": 4},
    {"key": "com_unit_name", "label": "单位", "show": True, "order": 5},
    {"key": "remark", "label": "备注", "show": False, "order": 6},
    {"key": "inv_class_name", "label": "分类名称", "show": False, "order": 7},
    {"key": "size_model", "label": "尺码型号", "show": False, "order": 8},
    {"key": "craft_desc", "label": "工艺", "show": False, "order": 9},
    {"key": "material_desc", "label": "材质", "show": False, "order": 10},
    {"key": "market_desc", "label": "市场区域", "show": False, "order": 11},
    {"key": "package_size", "label": "包装尺寸", "show": False, "order": 12},
]

PARAM_MAP = {
    "标签宽度": ("label_width_mm", float),
    "标签最小高度": ("label_min_height_mm", float),
    "内边距": ("label_padding_mm", float),
    "名称字体": ("name_font_family", str),
    "名称字号": ("name_font_size_pt", float),
    "名称最多行数": ("name_max_lines", int),
    "字段字体": ("field_font_family", str),
    "字段字号": ("field_font_size_pt", float),
    "备注最多行数": ("remark_max_lines", int),
    "标签列宽": ("field_label_width_px", int),
    "显示二维码": ("qr_show", None),
    "二维码尺寸": ("qr_size_mm", float),
    "每行标签数": ("labels_per_row", int),
    "标签间距": ("label_gap_mm", float),
    "每页标签数": ("labels_per_page", int),
    "显示底部信息": ("footer_show", None),
    "底部字体": ("footer_font_family", str),
    "底部字号": ("footer_font_size_pt", float),
    "底部左侧内容": ("footer_left_text", str),
    "底部右侧内容": ("footer_right_text", str),
}

BOOL_TRUE = {"是", "yes", "true", "1", "y"}
_cache: Dict[str, Tuple[float, Dict[str, Any]]] = {}
FONT_NAME_RE = re.compile(r"^[\w\u4e00-\u9fff -]+$")


def _to_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in BOOL_TRUE


def _defaults() -> Dict[str, Any]:
    cfg = DEFAULTS.copy()
    cfg["fields"] = [dict(field) for field in DEFAULT_FIELDS]
    cfg["visible_fields"] = [field for field in cfg["fields"] if field["show"]]
    cfg["field_label_width"] = f"{cfg['field_label_width_px']}px"
    return _normalize_config(cfg)


def _css_font_family(value: Any) -> str:
    parts = []
    for part in str(value or "").split(","):
        name = part.strip().strip("\"'")
        if name and FONT_NAME_RE.match(name):
            parts.append(name)
    return ", ".join(parts) or DEFAULTS["field_font_family"]


def _normalize_config(cfg: Dict[str, Any]) -> Dict[str, Any]:
    cfg["name_font_family"] = _css_font_family(cfg.get("name_font_family"))
    cfg["field_font_family"] = _css_font_family(cfg.get("field_font_family"))
    cfg["footer_font_family"] = _css_font_family(cfg.get("footer_font_family"))
    return cfg


def _read_excel(path: str) -> Dict[str, Any]:
    cfg = DEFAULTS.copy()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if "配置" in wb.sheetnames:
        for row in wb["配置"].iter_rows(min_row=2, values_only=True):
            if not row or row[1] is None:
                continue
            name = str(row[1]).strip()
            value = row[2] if len(row) > 2 else None
            if name not in PARAM_MAP or value is None:
                continue
            key, conv = PARAM_MAP[name]
            try:
                cfg[key] = _to_bool(value) if conv is None else conv(value)
            except (TypeError, ValueError):
                pass

    fields = [dict(field) for field in DEFAULT_FIELDS]
    index = {field["key"]: field for field in fields}
    if "字段配置" in wb.sheetnames:
        for row in wb["字段配置"].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            if key not in index:
                continue
            if len(row) > 1 and row[1]:
                index[key]["label"] = str(row[1]).strip()
            if len(row) > 2 and row[2] is not None:
                index[key]["show"] = _to_bool(row[2])
            if len(row) > 3 and row[3] is not None:
                try:
                    index[key]["order"] = int(row[3])
                except (TypeError, ValueError):
                    pass
    wb.close()

    cfg["fields"] = sorted(fields, key=lambda field: field["order"])
    cfg["visible_fields"] = [field for field in cfg["fields"] if field["show"]]
    cfg["field_label_width"] = f"{cfg['field_label_width_px']}px"
    return _normalize_config(cfg)


def get_label_config(module_cfg: ModuleConfig) -> Dict[str, Any]:
    path = resolve_module_path(module_cfg.template_path)
    try:
        mtime = os.path.getmtime(path)
    except FileNotFoundError:
        return _defaults()

    cached = _cache.get(path)
    if cached and cached[0] == mtime:
        return cached[1]

    try:
        data = _read_excel(path)
    except Exception:
        data = _defaults()
    _cache[path] = (mtime, data)
    return data
