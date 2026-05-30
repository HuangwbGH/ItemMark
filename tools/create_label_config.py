"""
生成初始 config/label_config.xlsx 文件。
可重复运行，会覆盖已有文件（重置为默认配置）。

用法：
    python tools/create_label_config.py
"""

import argparse
import os
import sys
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── 颜色常量 ─────────────────────────────────────────
BLUE_FILL   = PatternFill("solid", fgColor="1A6FC4")   # 表头背景
GRAY_FILL   = PatternFill("solid", fgColor="E8ECF0")   # 分组行背景
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL    = PatternFill("solid", fgColor="F7F9FB")   # 隔行背景

HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
GROUP_FONT  = Font(name="微软雅黑", bold=True, color="333333", size=10)
DATA_FONT   = Font(name="微软雅黑", size=10)
HINT_FONT   = Font(name="微软雅黑", size=9, color="888888", italic=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def apply_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = THIN_BORDER


def write_header(ws, row, cols):
    """写表头行"""
    for ci, text in enumerate(cols, start=1):
        cell = ws.cell(row, ci)
        cell.value      = text
        cell.font       = HEADER_FONT
        cell.fill       = BLUE_FILL
        cell.alignment  = CENTER
        cell.border     = THIN_BORDER


def write_group(ws, row, text, n_cols):
    """写分组标题行（合并单元格）"""
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cols)
    cell = ws.cell(row, 1)
    cell.value     = f"  {text}"
    cell.font      = GROUP_FONT
    cell.fill      = GRAY_FILL
    cell.alignment = LEFT
    cell.border    = THIN_BORDER
    for c in range(2, n_cols + 1):
        ws.cell(row, c).border = THIN_BORDER


def write_data_row(ws, row, values, alt=False):
    fill = ALT_FILL if alt else WHITE_FILL
    for ci, v in enumerate(values, start=1):
        cell = ws.cell(row, ci)
        cell.value     = v
        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.alignment = LEFT if ci > 2 else CENTER
        cell.border    = THIN_BORDER


# ── Sheet 1: 配置 ─────────────────────────────────────
CONFIG_ROWS = [
    # (分组, 参数名称, 当前值, 默认值, 说明, 单位)
    # 分组行用 None 作为分隔
    (None, "【标签尺寸】"),
    ("标签尺寸", "标签宽度",      100, 100, "标签纸宽度",                       "mm"),
    ("标签尺寸", "标签最小高度",   58,  58, "标签内容不足时的最小高度，内容超出时自动撑高", "mm"),
    ("标签尺寸", "内边距",          4,   4, "标签四边内部留白",                   "mm"),
    (None, "【名称区】"),
    ("名称区",   "名称字体", "Microsoft YaHei, PingFang SC, Arial, sans-serif", "Microsoft YaHei, PingFang SC, Arial, sans-serif", "顶部物料名称的字体，多个字体用英文逗号分隔", "-"),
    ("名称区",   "名称字号",       13,  13, "顶部物料名称的字体大小",             "pt"),
    ("名称区",   "名称最多行数",    2,   2, "超出行数时自动截断",                 "行"),
    (None, "【信息字段】"),
    ("信息字段", "字段字体", "Microsoft YaHei, PingFang SC, Arial, sans-serif", "Microsoft YaHei, PingFang SC, Arial, sans-serif", "各信息字段（编号/布卡号等）的字体，多个字体用英文逗号分隔", "-"),
    ("信息字段", "字段字号",       8.5, 8.5,"各信息字段（编号/布卡号等）的字体大小", "pt"),
    ("信息字段", "标签列宽",       52,  52, '字段标签文字（如"物料编号"）的列宽',  "px"),
    (None, "【二维码】"),
    ("二维码",   "显示二维码",     "是","是","是 / 否",                          "-"),
    ("二维码",   "二维码尺寸",     22,  22, "二维码图片边长",                     "mm"),
    (None, "【批量打印布局】"),
    ("批量打印", "每行标签数",      2,   2, "批量打印时每行排列几张标签",          "-"),
    ("批量打印", "标签间距",        6,   6, "相邻标签之间的间隙",                 "mm"),
    ("批量打印", "每页标签数",      4,   4, "超过此数量后自动换页（0 = 不限制）",  "-"),
    (None, "【底部信息】"),
    ("底部信息", "显示底部信息",   "是","是","是 / 否，控制标签底部打印时间等信息", "-"),
    ("底部信息", "底部字体", "Microsoft YaHei, PingFang SC, Arial, sans-serif", "Microsoft YaHei, PingFang SC, Arial, sans-serif", "底部编号与打印时间的字体，多个字体用英文逗号分隔", "-"),
    ("底部信息", "底部字号",        6,   6, "底部编号与打印时间的字体大小",        "pt"),
    ("底部信息", "底部左侧内容", "{material_code}", "{material_code}", "支持占位符：{material_code}、{material_name}、{print_time} 等", "-"),
    ("底部信息", "底部右侧内容", "{print_time}", "{print_time}", "支持占位符：{material_code}、{material_name}、{print_time} 等", "-"),
]

FIELD_ROWS = [
    # (字段键, 显示名称, 是否显示, 显示顺序)
    ("material_code",  "物料编号", "是", 1),
    ("cloth_card_no",  "布卡号",   "是", 2),
    ("specification",  "规格型号", "是", 3),
    ("color_desc",     "颜色",     "是", 4),
    ("com_unit_name",  "单位",     "是", 5),
    ("inv_class_name", "分类名称", "否", 6),
    ("size_model",     "尺码型号", "否", 7),
    ("craft_desc",     "工艺",     "否", 8),
    ("material_desc",  "材质",     "否", 9),
    ("market_desc",    "市场区域", "否", 10),
    ("package_size",   "包装尺寸", "否", 11),
]

MATERIAL_INFO_FIELD_ROWS = [
    ("material_code",  "物料编号", "是", 1),
    ("cloth_card_no",  "布卡号",   "是", 2),
    ("specification",  "规格型号", "是", 3),
    ("color_desc",     "颜色",     "是", 4),
    ("remark",         "备注",     "是", 5),
    ("inv_class_name", "分类名称", "否", 6),
    ("size_model",     "尺码型号", "否", 7),
    ("craft_desc",     "工艺",     "否", 8),
    ("material_desc",  "材质",     "否", 9),
    ("market_desc",    "市场区域", "否", 10),
    ("package_size",   "包装尺寸", "否", 11),
]


def build_config_sheet(wb):
    ws = wb.create_sheet("配置")
    ws.sheet_view.showGridLines = False

    # 说明行
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "物料标签打印配置 — 修改「当前值」列后保存，刷新打印页即自动生效"
    c.font      = Font(name="微软雅黑", bold=True, size=12, color="1A6FC4")
    c.alignment = LEFT
    ws.row_dimensions[1].height = 24

    # 表头
    write_header(ws, 2, ["参数分组", "参数名称", "当前值", "默认值", "说明", "单位"])
    ws.row_dimensions[2].height = 22

    row = 3
    alt = False
    for entry in CONFIG_ROWS:
        if entry[0] is None:
            # 分组标题行
            write_group(ws, row, entry[1], 6)
            ws.row_dimensions[row].height = 20
            row += 1
            alt = False
            continue

        group, name, cur_val, def_val, desc, unit = entry
        write_data_row(ws, row, [group, name, cur_val, def_val, desc, unit], alt=alt)
        # 「当前值」列加粗显示，提示用户这里可以修改
        ws.cell(row, 3).font = Font(name="微软雅黑", size=10, bold=True, color="1A6FC4")
        ws.row_dimensions[row].height = 18
        alt = not alt
        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 8

    # 冻结前两行
    ws.freeze_panes = "A3"


def build_field_sheet(wb, field_rows):
    ws = wb.create_sheet("字段配置")
    ws.sheet_view.showGridLines = False

    # 说明行
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = "控制标签上显示哪些字段及其顺序 — 「是否显示」填「是」或「否」，「显示顺序」数字越小越靠前"
    c.font      = Font(name="微软雅黑", bold=True, size=11, color="1A6FC4")
    c.alignment = LEFT
    ws.row_dimensions[1].height = 22

    write_header(ws, 2, ["字段键", "显示名称", "是否显示", "显示顺序"])
    ws.row_dimensions[2].height = 22

    for i, (key, label, show, order) in enumerate(field_rows, start=3):
        fill = ALT_FILL if i % 2 == 1 else WHITE_FILL
        for ci, v in enumerate([key, label, show, order], start=1):
            cell = ws.cell(i, ci)
            cell.value     = v
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.alignment = CENTER
            cell.border    = THIN_BORDER
        # 「是否显示」列标蓝
        ws.cell(i, 3).font = Font(name="微软雅黑", size=10, bold=True, color="1A6FC4")
        ws.row_dimensions[i].height = 18

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A3"


def main():
    parser = argparse.ArgumentParser(description="生成模块标签配置")
    parser.add_argument("--module", default="wuping", choices=["material-info", "wuping"])
    args = parser.parse_args()

    out_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "config", "templates", args.module,
    )
    out_path = os.path.join(out_dir, "label_config.xlsx")
    os.makedirs(out_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    build_config_sheet(wb)
    field_rows = MATERIAL_INFO_FIELD_ROWS if args.module == "material-info" else FIELD_ROWS
    build_field_sheet(wb, field_rows)

    wb.save(out_path)
    print(f"✓ 已生成：{out_path}")


if __name__ == "__main__":
    main()
